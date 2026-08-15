"""
Export builders: Excel (.xlsx) and PDF (via WeasyPrint HTML→PDF).

Two export modes:
  - per_class  : all students, all components, avg, freq, status for one class
  - full_system: all classes in one workbook (one sheet per ClassOffering)

PDF export renders a Jinja2 template to HTML, then converts with WeasyPrint.
WeasyPrint requires system libs (pango, cairo) on Linux — see deploy/setup.sh.
On Windows dev machines without WeasyPrint, the xlsx-only path still works.
"""

import io
import os
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import current_app, render_template

from app.utils import (
    compute_student_average,
    compute_student_freq,
    compute_student_status,
    build_grade_map,
    format_grade,
    STATUS_LABELS,
)


# ─────────────────────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────────────────────

_HEADER_FILL    = PatternFill("solid", fgColor="0D1B2A")
_HEADER_FONT    = Font(bold=True, color="FFFFFF", name="Calibri")
_ALT_FILL       = PatternFill("solid", fgColor="F0F4F8")
_CENTER         = Alignment(horizontal="center", vertical="center")
_THIN           = Side(style="thin", color="CCCCCC")
_BORDER         = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_header_row(ws, row_num: int, ncols: int):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER


def export_class_xlsx(
    class_offering,
    columns: list[str] | None = None,
    anonymized: bool = False,
) -> bytes:
    """
    Generate an Excel workbook for one class.

    columns: list of column keys to include. None = all.
    anonymized: if True, omit Nome column.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _safe_sheet_name(class_offering)

    components = [c for c in class_offering.components if not c.is_ps]
    ps_comps   = [c for c in class_offering.components if c.is_ps]
    all_comps  = class_offering.components

    # Build header
    headers = ["Seq", "Cartão"]
    if not anonymized:
        headers.append("Nome")
    for comp in all_comps:
        headers.append(comp.name)
    headers += ["Média", "Frequência %", "Situação"]

    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)
    _style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 20

    # Data rows
    enrollments = sorted(
        class_offering.active_enrollments,
        key=lambda e: e.seq or 0,
    )
    settings = class_offering.settings

    for row_num, enrollment in enumerate(enrollments, 2):
        gmap    = build_grade_map(enrollment)
        records = enrollment.attendance_records
        avg     = compute_student_average(enrollment, all_comps, gmap)
        freq    = compute_student_freq(records)
        status  = compute_student_status(avg, freq, settings)

        col = 1
        ws.cell(row=row_num, column=col, value=enrollment.seq);          col += 1
        ws.cell(row=row_num, column=col, value=enrollment.student.cartao); col += 1
        if not anonymized:
            ws.cell(row=row_num, column=col, value=enrollment.student.nome); col += 1

        for comp in all_comps:
            g = gmap.get(comp.id)
            if g is None:
                val = ""
            elif g.status == "missed_unjustified":
                val = "FI"
            elif g.status == "missed_justified":
                val = "FJ"
            else:
                val = g.value if g.value is not None else ""
            ws.cell(row=row_num, column=col, value=val)
            col += 1

        ws.cell(row=row_num, column=col, value=round(avg, 2) if avg is not None else ""); col += 1
        ws.cell(row=row_num, column=col,
                value=f"{freq:.1f}%" if freq is not None else "—"); col += 1
        ws.cell(row=row_num, column=col,
                value=STATUS_LABELS.get(status, status)); col += 1

        if row_num % 2 == 0:
            for c in range(1, col):
                ws.cell(row=row_num, column=c).fill = _ALT_FILL

    # Auto-width
    for col_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(length + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_full_system_xlsx(all_classes) -> bytes:
    """One workbook, one sheet per ClassOffering."""
    wb = openpyxl.Workbook()
    first = True

    for co in all_classes:
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = _safe_sheet_name(co)

        components = co.components
        headers = ["Seq", "Cartão", "Nome"] + [c.name for c in components] + ["Média", "Freq%", "Situação"]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)
        _style_header_row(ws, 1, len(headers))

        for row_num, enrollment in enumerate(sorted(co.active_enrollments, key=lambda e: e.seq or 0), 2):
            gmap = build_grade_map(enrollment)
            avg  = compute_student_average(enrollment, components, gmap)
            freq = compute_student_freq(enrollment.attendance_records)
            status = compute_student_status(avg, freq, co.settings)

            col = 1
            ws.cell(row=row_num, column=col, value=enrollment.seq);               col += 1
            ws.cell(row=row_num, column=col, value=enrollment.student.cartao);    col += 1
            ws.cell(row=row_num, column=col, value=enrollment.student.nome);      col += 1
            for comp in components:
                g = gmap.get(comp.id)
                ws.cell(row=row_num, column=col,
                        value="" if g is None else (g.value if g.status == "graded" else g.status)); col += 1
            ws.cell(row=row_num, column=col, value=round(avg, 2) if avg is not None else ""); col += 1
            ws.cell(row=row_num, column=col, value=f"{freq:.1f}" if freq is not None else ""); col += 1
            ws.cell(row=row_num, column=col, value=STATUS_LABELS.get(status, "")); col += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _safe_sheet_name(class_offering) -> str:
    name = f"{class_offering.subject_name[:20]} {class_offering.turma} {class_offering.periodo_letivo}"
    # Excel sheet names: max 31 chars, no special chars
    for ch in r"\/*?[]":
        name = name.replace(ch, "")
    return name[:31]


# ─────────────────────────────────────────────────────────────────────────────
# PDF export (WeasyPrint HTML→PDF)
# ─────────────────────────────────────────────────────────────────────────────

def export_class_pdf(
    class_offering,
    anonymized: bool = False,
    failing_only: bool = False,
    student_id: int | None = None,
) -> bytes:
    """
    Generate a PDF for one class.

    Renders templates/export/class_pdf.html, then passes through WeasyPrint.
    Falls back to a plain text message if WeasyPrint is not installed.
    """
    try:
        from weasyprint import HTML
    except ImportError:
        raise RuntimeError(
            "WeasyPrint não está instalado. Execute: pip install weasyprint\n"
            "No Debian também é necessário: sudo apt install libpango-1.0-0 libpangoft2-1.0-0"
        )

    components = class_offering.components
    settings   = class_offering.settings

    enrollments = class_offering.active_enrollments
    if student_id:
        enrollments = [e for e in enrollments if e.student_id == student_id]

    rows = []
    for enrollment in sorted(enrollments, key=lambda e: e.seq or 0):
        gmap    = build_grade_map(enrollment)
        records = enrollment.attendance_records
        avg     = compute_student_average(enrollment, components, gmap)
        freq    = compute_student_freq(records)
        status  = compute_student_status(avg, freq, settings)

        if failing_only and status not in ("reprovado_nota", "reprovado_falta", "reprovado_ambos"):
            continue

        rows.append({
            "enrollment": enrollment,
            "gmap":       gmap,
            "avg":        avg,
            "freq":       freq,
            "status":     status,
        })

    html_content = render_template(
        "export/class_pdf.html",
        class_offering=class_offering,
        components=components,
        rows=rows,
        anonymized=anonymized,
        failing_only=failing_only,
        generated_at=datetime.now(timezone.utc),
    )

    return HTML(string=html_content).write_pdf()
